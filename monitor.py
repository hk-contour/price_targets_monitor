#!/usr/bin/env python3
"""
Contour Price Target Monitor
- Reads Excel from OneDrive via a share link (no Azure app needed)
- Checks live prices via yfinance every 4 hours
- Emails via Outlook SMTP when price is within 10% of a target
- One alert per ticker per calendar day
"""

import os
import io
import json
import smtplib
import logging
import time
import traceback
from datetime import datetime, date
from collections import defaultdict
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import yfinance as yf
import requests
import openpyxl

# ─────────────────────────────────────────────────────────────
# CONFIGURATION  — fill these in (or set as env vars / GitHub Secrets)
# ─────────────────────────────────────────────────────────────

# OneDrive/SharePoint share link for the Excel file.
# In SharePoint: open the file → Share → Copy Link → set to "Anyone with link can view"
# Paste that URL here. The script converts it to a direct download automatically.
SHAREPOINT_SHARE_URL = os.getenv(
    "SHAREPOINT_SHARE_URL",
    "https://contourassetmgmt-my.sharepoint.com/:x:/r/personal/hari_kumar_contourasset_com/_layouts/15/Doc.aspx?sourcedoc=%7BF34FAAAE-2AA6-4132-B9D3-4B8C5F416DFA%7D&file=Contour-Price-Targets.xlsx&action=default&mobileredirect=true"
)

# Outlook SMTP — use an App Password (see SETUP.md)
SMTP_HOST   = "smtp.office365.com"
SMTP_PORT   = 587
SMTP_USER   = os.getenv("SMTP_USER", "hari.kumar@contourasset.com")  # sending account
SMTP_PASS   = os.getenv("SMTP_PASS", "")                             # App Password

ALERT_EMAILS = [
    "hari.kumar@contourasset.com",
    "david.meyer@contourasset.com",
    "Michael.tabaksblat@contourasset.com",
]

THRESHOLD_PCT  = 0.10   # 10 % proximity
ALERT_LOG_FILE = "alerts_sent.json"

# ─────────────────────────────────────────────────────────────
# LOGGING
# ─────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler("monitor.log"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────
# STEP 1: DOWNLOAD EXCEL FROM ONEDRIVE
# ─────────────────────────────────────────────────────────────

def sharepoint_share_url_to_download_url(share_url: str) -> str:
    """
    Converts a SharePoint/OneDrive share URL into a direct-download URL
    via the OneDrive API — no auth required if the link is set to
    'Anyone with the link can view'.
    """
    import base64
    # Encode the share URL as base64url (OneDrive sharing API format)
    b64 = base64.urlsafe_b64encode(share_url.encode()).decode().rstrip("=")
    sharing_token = f"u!{b64}"
    api_url = f"https://api.onedrive.com/v1.0/shares/{sharing_token}/root/content"
    return api_url


def download_excel_bytes() -> bytes:
    """Download the Excel file and return its raw bytes."""
    dl_url = sharepoint_share_url_to_download_url(SHAREPOINT_SHARE_URL)
    log.info(f"Downloading Excel from OneDrive…")
    resp = requests.get(dl_url, timeout=30, allow_redirects=True)

    if resp.status_code == 401:
        raise RuntimeError(
            "OneDrive returned 401 Unauthorized.\n"
            "Make sure the SharePoint share link is set to 'Anyone with the link can view'.\n"
            "See SETUP.md for instructions."
        )
    if resp.status_code != 200:
        raise RuntimeError(
            f"Failed to download Excel (HTTP {resp.status_code}).\n"
            f"URL tried: {dl_url}"
        )
    log.info(f"Downloaded {len(resp.content):,} bytes.")
    return resp.content


# ─────────────────────────────────────────────────────────────
# STEP 2: PARSE EXCEL → TARGETS
# ─────────────────────────────────────────────────────────────

def parse_excel(excel_bytes: bytes) -> dict:
    """
    Parse the workbook and return:
      { ticker: { upside, downside, date } }
    using the most-recent BeginDate row per ticker that has ≥1 target.
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    # ── Find header row ──────────────────────────────────────
    col = {}   # logical name → 0-based column index
    header_row_idx = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any(k in cells for k in ("issuer", "ticker")):
            for j, h in enumerate(cells):
                if h in ("issuer", "ticker"):
                    col["ticker"] = j
                elif "upside" in h:
                    col["upside"] = j
                elif "downside" in h:
                    col["downside"] = j
                elif "begin" in h or (h == "date" and "begin" not in col):
                    col.setdefault("begin_date", j)
                elif "end" in h:
                    col.setdefault("end_date", j)
            header_row_idx = i
            log.info(f"Header found at row {i+1}: {col}")
            break

    if not col or "ticker" not in col:
        raise RuntimeError(
            "Could not find a header row with 'Issuer'/'Ticker' in the Excel file. "
            "Check that the spreadsheet format hasn't changed."
        )

    # ── Parse data rows ──────────────────────────────────────
    rows = []
    skipped_no_date = []

    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        def get(key):
            idx = col.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        ticker = get("ticker")
        if not ticker:
            continue
        ticker = str(ticker).strip().upper()
        if not ticker or ticker in ("ISSUER", "TICKER", "NONE", "NAN", ""):
            continue

        # Parse BeginDate
        raw_date = get("begin_date")
        entry_date = None
        if isinstance(raw_date, (datetime, date)):
            entry_date = raw_date.date() if isinstance(raw_date, datetime) else raw_date
        elif isinstance(raw_date, str) and raw_date.strip():
            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%d/%m/%Y"):
                try:
                    entry_date = datetime.strptime(raw_date.strip(), fmt).date()
                    break
                except ValueError:
                    pass
        elif isinstance(raw_date, (int, float)):
            # Excel serial date
            try:
                from openpyxl.utils.datetime import from_excel
                entry_date = from_excel(raw_date)
                if isinstance(entry_date, datetime):
                    entry_date = entry_date.date()
            except Exception:
                pass

        if entry_date is None:
            skipped_no_date.append(ticker)
            continue

        def to_float(v):
            try:
                return float(v) if v is not None else None
            except (ValueError, TypeError):
                return None

        rows.append({
            "date":     entry_date,
            "ticker":   ticker,
            "upside":   to_float(get("upside")),
            "downside": to_float(get("downside")),
        })

    if skipped_no_date:
        log.warning(f"Skipped {len(skipped_no_date)} rows with no parseable date (sample: {skipped_no_date[:5]})")

    log.info(f"Parsed {len(rows)} data rows, {len(set(r['ticker'] for r in rows))} unique tickers.")

    # ── Pick most-recent row per ticker ──────────────────────
    by_ticker = defaultdict(list)
    for r in rows:
        by_ticker[r["ticker"]].append(r)

    targets = {}
    no_targets = []
    for ticker, entries in sorted(by_ticker.items()):
        sorted_entries = sorted(entries, key=lambda x: x["date"], reverse=True)
        chosen = next(
            (e for e in sorted_entries if e["upside"] is not None or e["downside"] is not None),
            None,
        )
        if chosen is None:
            no_targets.append(ticker)
            continue
        targets[ticker] = {
            "upside":   chosen["upside"],
            "downside": chosen["downside"],
            "date":     chosen["date"].isoformat(),
        }

    if no_targets:
        log.warning(f"No usable targets for {len(no_targets)} tickers: {no_targets}")

    log.info(f"Final: {len(targets)} tickers with targets.")
    return targets


# ─────────────────────────────────────────────────────────────
# STEP 3: LIVE PRICE FETCHING
# ─────────────────────────────────────────────────────────────

# Tickers in the spreadsheet that need exchange suffixes for yfinance
TICKER_MAP = {
    # Japan (TSE)
    "8136": "8136.T", "6098": "6098.T", "7974": "7974.T", "7751": "7751.T",
    "4324": "4324.T", "2330": "2330.T", "6981": "6981.T", "6963": "6963.T",
    "6857": "6857.T", "4661": "4661.T", "6594": "6594.T",
    # Germany (Xetra)
    "IFXGn": "IFX.DE", "AG1G": "AG1.DE", "SAPG": "SAP.DE",
    # UK
    "WISEa": "WISE.L", "PUBP": "PUB.L",
    # Canada
    "RCIb": "RCI-B.TO",
}

# Tickers that cannot be priced via yfinance (OTC-only, delisted, or non-standard)
UNTRADEABLE = {
    "MSCHWCCH", "MSCHWCHK", "FLTRF", "TEPRF", "LOOMb", "PRSM LN", "JET LN",
    "SIM", "WLN", "ASOS", "AUTOA", "RMV", "TKWY", "HFGG", "DHER",
    "ITRK", "PSON", "TMV", "BCO", "CHYM", "GRND", "MNTN",
}


def fetch_price(ticker: str):
    """
    Returns (price: float, currency: str) or (None, None).
    Tries fast_info first, falls back to 2-day history.
    Logs clearly on failure — nothing is silently dropped.
    """
    if ticker in UNTRADEABLE:
        log.debug(f"{ticker}: known untradeable, skipping.")
        return None, None

    yf_sym = TICKER_MAP.get(ticker, ticker)

    try:
        t    = yf.Ticker(yf_sym)
        fi   = t.fast_info
        price = getattr(fi, "last_price", None)

        if price is None or price <= 0:
            hist = t.history(period="2d", auto_adjust=True)
            if hist.empty:
                log.warning(f"{ticker} ({yf_sym}): no history data returned.")
                return None, None
            price = float(hist["Close"].iloc[-1])

        currency = getattr(fi, "currency", "USD") or "USD"
        return float(price), str(currency)

    except Exception as e:
        log.warning(f"{ticker} ({yf_sym}): price fetch failed — {type(e).__name__}: {e}")
        return None, None


# ─────────────────────────────────────────────────────────────
# STEP 4: DAILY ALERT DEDUPLICATION
# ─────────────────────────────────────────────────────────────

def load_alert_log() -> dict:
    if os.path.exists(ALERT_LOG_FILE):
        with open(ALERT_LOG_FILE) as f:
            return json.load(f)
    return {}


def save_alert_log(data: dict):
    with open(ALERT_LOG_FILE, "w") as f:
        json.dump(data, f, indent=2)


def already_alerted(alert_log: dict, ticker: str) -> bool:
    return alert_log.get(date.today().isoformat(), {}).get(ticker, False)


def mark_alerted(alert_log: dict, ticker: str):
    today = date.today().isoformat()
    alert_log.setdefault(today, {})[ticker] = True
    # Prune entries older than 14 days
    cutoff = date.today().toordinal() - 14
    for k in [k for k in alert_log if date.fromisoformat(k).toordinal() < cutoff]:
        del alert_log[k]


# ─────────────────────────────────────────────────────────────
# STEP 5: EMAIL
# ─────────────────────────────────────────────────────────────

def send_email(alerts: list):
    """
    alerts: list of dicts with ticker, price, currency,
            target_type, target_price, pct_away, target_date
    """
    if not SMTP_PASS:
        log.error("SMTP_PASS not set — cannot send email. Set it as an env var.")
        return

    n     = len(alerts)
    today = date.today().strftime("%B %d, %Y")
    subj  = f"[Contour Alert] {n} ticker{'s' if n>1 else ''} near price target — {today}"

    # ── HTML ──────────────────────────────────────────────────
    rows_html = ""
    for a in sorted(alerts, key=lambda x: x["pct_away"]):
        color     = "#c0392b" if a["pct_away"] < 3 else "#e67e22" if a["pct_away"] < 7 else "#2c3e50"
        direction = "▲ Upside" if a["target_type"] == "upside" else "▼ Downside"
        rows_html += f"""
          <tr>
            <td style="padding:8px 14px;font-weight:600;font-size:15px">{a['ticker']}</td>
            <td style="padding:8px 14px">{a['currency']} {a['price']:.2f}</td>
            <td style="padding:8px 14px">{direction}</td>
            <td style="padding:8px 14px">{a['currency']} {a['target_price']:.2f}</td>
            <td style="padding:8px 14px;font-weight:700;color:{color}">{a['pct_away']:.1f}% away</td>
            <td style="padding:8px 14px;color:#888;font-size:12px">{a['target_date']}</td>
          </tr>"""

    html = f"""
    <html><body style="font-family:'Segoe UI',Arial,sans-serif;background:#f5f5f5;margin:0;padding:24px">
      <div style="max-width:700px;margin:auto;background:white;border-radius:8px;
                  box-shadow:0 2px 8px rgba(0,0,0,0.08);overflow:hidden">
        <div style="background:#1a3c6e;padding:20px 28px">
          <h2 style="margin:0;color:white;font-size:20px">Contour Price Target Alert</h2>
          <p style="margin:4px 0 0;color:#a8c4e8;font-size:13px">{today}</p>
        </div>
        <div style="padding:24px 28px">
          <p style="margin:0 0 16px;color:#333">
            {n} ticker{'s are' if n>1 else ' is'} within <strong>{int(THRESHOLD_PCT*100)}%</strong>
            of a price target:
          </p>
          <table border="0" cellspacing="0" cellpadding="0"
                 style="width:100%;border-collapse:collapse;border:1px solid #e0e0e0;border-radius:6px;overflow:hidden">
            <thead>
              <tr style="background:#f0f4fa;color:#444;font-size:12px;text-transform:uppercase;letter-spacing:.5px">
                <th style="padding:8px 14px;text-align:left">Ticker</th>
                <th style="padding:8px 14px;text-align:left">Live Price</th>
                <th style="padding:8px 14px;text-align:left">Target Type</th>
                <th style="padding:8px 14px;text-align:left">Target</th>
                <th style="padding:8px 14px;text-align:left">Distance</th>
                <th style="padding:8px 14px;text-align:left">Set On</th>
              </tr>
            </thead>
            <tbody>{rows_html}</tbody>
          </table>
          <p style="margin:20px 0 0;font-size:11px;color:#aaa">
            Alerts fire at most once per ticker per calendar day · 
            Source: Contour-Price-Targets.xlsx
          </p>
        </div>
      </div>
    </body></html>"""

    # ── Plain text fallback ───────────────────────────────────
    text_rows = "\n".join(
        f"  {a['ticker']:8s}  live={a['price']:.2f}  "
        f"{'upside' if a['target_type']=='upside' else 'downside'}={a['target_price']:.2f}  "
        f"{a['pct_away']:.1f}% away  (set {a['target_date']})"
        for a in sorted(alerts, key=lambda x: x["pct_away"])
    )
    text = f"Contour Price Target Alert — {today}\n\n{n} ticker(s) within {int(THRESHOLD_PCT*100)}%:\n\n{text_rows}\n"

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subj
    msg["From"]    = SMTP_USER
    msg["To"]      = ", ".join(ALERT_EMAILS)
    msg.attach(MIMEText(text, "plain"))
    msg.attach(MIMEText(html,  "html"))

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as srv:
            srv.ehlo()
            srv.starttls()
            srv.login(SMTP_USER, SMTP_PASS)
            srv.sendmail(SMTP_USER, ALERT_EMAILS, msg.as_string())
        log.info(f"Email sent: {[a['ticker'] for a in alerts]}")
    except smtplib.SMTPAuthenticationError:
        log.error(
            "SMTP authentication failed. If MFA is on, you need an App Password.\n"
            "See SETUP.md → Step 2."
        )
    except Exception as e:
        log.error(f"Email failed: {e}")


# ─────────────────────────────────────────────────────────────
# MAIN CHECK
# ─────────────────────────────────────────────────────────────

def run_check():
    log.info("=" * 60)
    log.info("Starting price target check")

    # 1. Get targets
    try:
        excel_bytes = download_excel_bytes()
        targets     = parse_excel(excel_bytes)
    except Exception as e:
        log.error(f"FATAL: Could not load targets — {e}\n{traceback.format_exc()}")
        return

    # 2. Load dedup log
    alert_log = load_alert_log()

    # 3. Check prices
    alerts_to_send = []
    price_errors   = []

    for ticker, info in sorted(targets.items()):
        upside   = info.get("upside")
        downside = info.get("downside")
        tgt_date = info.get("date", "")

        if upside is None and downside is None:
            continue
        if already_alerted(alert_log, ticker):
            log.debug(f"{ticker}: already alerted today.")
            continue

        price, currency = fetch_price(ticker)

        if price is None:
            if ticker not in UNTRADEABLE:
                price_errors.append(ticker)
            continue

        triggered = False

        if upside is not None:
            pct = abs(price - upside) / upside * 100
            if pct <= THRESHOLD_PCT * 100:
                log.info(f"ALERT  {ticker}: ${price:.2f}  upside=${upside}  {pct:.1f}% away")
                alerts_to_send.append({
                    "ticker": ticker, "price": price, "currency": currency,
                    "target_type": "upside", "target_price": upside,
                    "pct_away": pct, "target_date": tgt_date,
                })
                triggered = True

        if downside is not None:
            pct = abs(price - downside) / downside * 100
            if pct <= THRESHOLD_PCT * 100:
                log.info(f"ALERT  {ticker}: ${price:.2f}  downside=${downside}  {pct:.1f}% away")
                alerts_to_send.append({
                    "ticker": ticker, "price": price, "currency": currency,
                    "target_type": "downside", "target_price": downside,
                    "pct_away": pct, "target_date": tgt_date,
                })
                triggered = True

        if triggered:
            mark_alerted(alert_log, ticker)

    save_alert_log(alert_log)

    # 4. Send email
    if alerts_to_send:
        send_email(alerts_to_send)
    else:
        log.info("No alerts this run.")

    # 5. Summary
    if price_errors:
        log.warning(f"Could not fetch prices for {len(price_errors)} tickers: {price_errors}")
    log.info(
        f"Done. {len(targets)} tickers checked, "
        f"{len(alerts_to_send)} alert(s), "
        f"{len(price_errors)} price error(s)."
    )


# ─────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse, schedule
    parser = argparse.ArgumentParser()
    parser.add_argument("--once", action="store_true", help="Run once and exit (good for testing)")
    args = parser.parse_args()

    if args.once:
        run_check()
    else:
        import schedule
        log.info("Scheduler started — running every 4 hours.")
        run_check()  # immediate first run
        schedule.every(4).hours.do(run_check)
        while True:
            schedule.run_pending()
            time.sleep(60)
